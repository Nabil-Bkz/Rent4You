"""
Export service for generating PDF and Excel files
"""
import io
from typing import List, Dict, Any
from django.http import HttpResponse
from django.db.models import QuerySet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch


class ExportService:
    """Service for exporting data to various formats"""
    
    @staticmethod
    def export_reservations_to_excel(reservations: QuerySet, filename: str = 'reservations.xlsx') -> HttpResponse:
        """
        Export reservations to Excel
        
        Args:
            reservations: QuerySet of reservations
            filename: Output filename
            
        Returns:
            HttpResponse with Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Réservations"
        
        # Headers
        headers = ['ID', 'Véhicule', 'Locataire', 'Date début', 'Date fin', 'Prix', 'Statut', 'Date création']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        for reservation in reservations:
            ws.append([
                reservation.id,
                f"{reservation.vehicule.marque} {reservation.vehicule.model}",
                reservation.locataire.user.email,
                reservation.date_debut.strftime('%Y-%m-%d'),
                reservation.date_fin.strftime('%Y-%m-%d'),
                float(reservation.prix),
                reservation.get_status_display(),
                reservation.created_at.strftime('%Y-%m-%d %H:%M')
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @staticmethod
    def export_reservations_to_pdf(reservations: QuerySet, filename: str = 'reservations.pdf') -> HttpResponse:
        """
        Export reservations to PDF
        
        Args:
            reservations: QuerySet of reservations
            filename: Output filename
            
        Returns:
            HttpResponse with PDF file
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#130f40'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Title
        title = Paragraph("Liste des Réservations", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Table data
        data = [['ID', 'Véhicule', 'Locataire', 'Date début', 'Date fin', 'Prix', 'Statut']]
        
        for reservation in reservations:
            data.append([
                str(reservation.id),
                f"{reservation.vehicule.marque} {reservation.vehicule.model}",
                reservation.locataire.user.email,
                reservation.date_debut.strftime('%Y-%m-%d'),
                reservation.date_fin.strftime('%Y-%m-%d'),
                f"{reservation.prix} MAD",
                reservation.get_status_display(),
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ff7800')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @staticmethod
    def export_vehicles_to_excel(vehicles: QuerySet, filename: str = 'vehicles.xlsx') -> HttpResponse:
        """Export vehicles to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Véhicules"
        
        headers = ['ID', 'Matricule', 'Marque', 'Modèle', 'Catégorie', 'Prix/jour', 'Prix/heure', 'État', 'Disponibilité']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        for vehicle in vehicles:
            ws.append([
                vehicle.id,
                vehicle.matricule,
                vehicle.marque,
                vehicle.model,
                vehicle.get_categorie_vehicule_display(),
                float(vehicle.prix_jour),
                float(vehicle.prix_heure),
                vehicle.get_etat_vehicule_display(),
                'Oui' if vehicle.disponibilite else 'Non',
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

